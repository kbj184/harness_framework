"""config 주도 제너릭 파서 — xlsx 14시트 → LedgerAsset (DB 미적재)."""

from __future__ import annotations

import logging
from collections import defaultdict

from openpyxl import load_workbook

from src.agents.excel_ledger import normalize as nz
from src.agents.excel_ledger.config import CPE_SOURCE, SHEET_CODE, SHEET_SPECS, SheetSpec
from src.agents.excel_ledger.cpe import cpe_for
from src.agents.excel_ledger.hashing import asset_hash
from src.agents.excel_ledger.models import LedgerAsset, SheetStat

logger = logging.getLogger("collect_cmdb")

# LedgerAsset 정형 컬럼으로 소비되는 필드 (나머지 spec.fields → attributes)
_STRUCTURED = {
    "hostname", "primary_ip", "os_name", "os_version", "manufacturer", "model",
    "isms", "c", "i", "a", "grade", "owner", "dept", "no", "mac",
}
# natural_key 매칭 시 소문자화할 필드(호스트류)
_LOWER_FIELDS = {"hostname", "cn", "uid", "dom"}


class HeaderMismatchError(ValueError):
    """헤더 assert 실패 — 컬럼 인덱스 드리프트."""


def _norm_value(field: str, raw_fields: dict[str, str]) -> str:
    """natural_key/식별용 정규화 값."""
    if field == "account_key":
        return nz.account_key(raw_fields.get("account_raw", ""))
    val = raw_fields.get(field, "")
    if field == "dom":  # 도메인은 선택 필드 — 공백은 'NA'로 coalesce (§8.6), 키 결측 아님
        return val.lower() or "na"
    if field in _LOWER_FIELDS:
        return val.lower()
    if field == "primary_ip":
        return nz.ip_primary(val)[0] or ""
    return val


def _build_norm(raw_fields: dict[str, str], spec: SheetSpec) -> dict[str, str]:
    """natural_key가 참조하는 필드 + account_key 정규화 값 dict."""
    norm = {f: _norm_value(f, raw_fields) for f in spec.natural_key}
    if "account_raw" in raw_fields:
        norm["account_key"] = nz.account_key(raw_fields["account_raw"])
    return norm


def _assert_headers(ws, spec: SheetSpec) -> dict[int, str]:
    """헤더 행 읽어 expected substring assert. 컬럼→헤더명 반환."""
    header_cells = {}
    for row in ws.iter_rows(min_row=spec.header_row, max_row=spec.header_row, values_only=True):
        for ci, v in enumerate(row, start=1):
            header_cells[ci] = nz.clean(v)
        break
    for col, expect in spec.asserts.items():
        actual = header_cells.get(col, "")
        if expect not in actual:
            raise HeaderMismatchError(
                f"[{spec.sheet_name}] col{col} 기대='{expect}' 실제='{actual}' — 컬럼 드리프트"
            )
    return header_cells


def parse_sheet(ws, spec: SheetSpec) -> tuple[list[LedgerAsset], SheetStat]:
    headers = _assert_headers(ws, spec)

    parsed: list[tuple[dict, dict, dict]] = []  # (raw_fields, norm, raw_row)
    for row in ws.iter_rows(min_row=spec.header_row + 1, max_row=ws.max_row, values_only=True):
        no = nz.clean(row[spec.fields["no"] - 1]) if len(row) >= spec.fields["no"] else ""
        if not no:
            continue
        raw_fields = {
            f: (nz.clean(row[idx - 1]) if len(row) >= idx else "")
            for f, idx in spec.fields.items()
        }
        raw_row = {
            (headers.get(ci) or f"c{ci}"): nz.clean(v)
            for ci, v in enumerate(row, start=1) if nz.clean(v)
        }
        parsed.append((raw_fields, _build_norm(raw_fields, spec), raw_row))

    # 충돌 패스: natural_key 그룹핑
    groups: dict[tuple, list[int]] = defaultdict(list)
    for i, (_, norm, _) in enumerate(parsed):
        groups[tuple(norm[f] for f in spec.natural_key)].append(i)

    assets: list[LedgerAsset] = []
    review_breakdown: dict[str, int] = defaultdict(int)
    for rf, norm, raw_row in parsed:
        key = tuple(norm[f] for f in spec.natural_key)
        parts = list(key)
        review_flag = None
        use_no = None

        if any(p == "" for p in parts):  # 약한 키(결측·TBD)
            use_no = rf["no"]
            ip_missing = "primary_ip" in spec.natural_key and norm.get("primary_ip", "") == ""
            review_flag = "TBD_IP" if ip_missing else "MISSING_KEY"
        elif len(groups[key]) > 1:  # 시트내 충돌
            use_no = rf["no"]
            review_flag = spec.collision_flag

        asset = _to_asset(rf, norm, raw_row, spec, parts, use_no, review_flag)
        assets.append(asset)
        if review_flag:
            review_breakdown[review_flag] += 1

    if spec.review_on_dup_hostname:
        _flag_dup_hostname_same_ip(parsed, assets, review_breakdown)

    stat = SheetStat(
        sheet=spec.sheet_name,
        category_cd=spec.category_cd,
        rows=len(assets),
        unique_hash=len({a.asset_id_hash for a in assets}),
        auto=sum(1 for a in assets if a.review_flag is None),
        review_queue=sum(1 for a in assets if a.review_flag is not None),
        review_breakdown=dict(review_breakdown),
    )
    return assets, stat


def _to_asset(rf, norm, raw_row, spec, parts, use_no, review_flag) -> LedgerAsset:
    primary_ip, ip_list = nz.ip_primary(rf.get("primary_ip", ""))
    pub = nz.clean(rf.get("pub_ip", ""))
    if pub:
        _, pubs = nz.ip_primary(pub)
        ip_list = ip_list + [p for p in pubs if p not in ip_list]

    c, i, a = nz.to_int(rf.get("c")), nz.to_int(rf.get("i")), nz.to_int(rf.get("a"))
    score = (c + i + a) if None not in (c, i, a) else None

    attributes = {
        f: v for f, v in rf.items()
        if f not in _STRUCTURED and v
    }
    if "account_raw" in rf and rf["account_raw"]:
        attributes["account_key"] = norm.get("account_key")
        attributes["account_provider"] = nz.account_provider(norm.get("account_key", ""))

    macs = []
    mac = nz.clean(rf.get("mac", ""))
    if mac:
        macs = [mac]

    # CPE 매핑 (§10.1 D4) — 카테고리별 product/version 원천
    cpe = None
    pf, vf = CPE_SOURCE.get(spec.category_cd, (None, None))
    if pf:
        cpe = cpe_for(rf.get(pf, ""), rf.get(vf, "") if vf else "")
        if cpe.unmapped:
            attributes["cpe_unmapped"] = True
            attributes["cpe_tier"] = cpe.tier

    return LedgerAsset(
        source_id=f"{SHEET_CODE[spec.sheet_name]}-{rf['no']}",
        sheet=spec.sheet_name,
        category_cd=spec.category_cd,
        asset_id_hash=asset_hash(spec.category_cd, parts, use_no),
        hostname=nz.clean(rf.get("hostname")) or None,
        primary_ip=primary_ip,
        ip_addresses=ip_list,
        os_name=nz.clean(rf.get("os_name")) or None,
        os_version=nz.clean(rf.get("os_version")) or None,
        manufacturer=nz.clean(rf.get("manufacturer")) or None,
        model=nz.clean(rf.get("model")) or None,
        isms_yn=nz.isms_yn(rf.get("isms")),
        confidentiality=c, integrity=i, availability=a,
        criticality_score=score,
        criticality_grade=nz.clean(rf.get("grade")) or None,
        env_type=nz.env_type(rf.get("location"), spec.env_default),
        owner_user_nm=nz.clean(rf.get("owner")) or None,
        owner_dept=nz.clean(rf.get("dept")) or None,
        cpe_uri=cpe.cpe_uri if cpe else None,
        cpe_vendor=cpe.vendor if cpe else None,
        cpe_product=cpe.product if cpe else None,
        cpe_version=cpe.version if cpe else None,
        cpe_tier=cpe.tier if cpe else None,
        lifecycle_state="CANDIDATE" if use_no else "ACTIVE",
        mac_addresses=macs,
        attributes=attributes,
        raw_data=raw_row,
        review_flag=review_flag,
    )


def _flag_dup_hostname_same_ip(parsed, assets, review_breakdown) -> None:
    """동명 호스트 중 primary_ip까지 동일한 행 = 1대 2뷰 의심 → 의미검토(PK는 유지)."""
    by_host: dict[str, list[int]] = defaultdict(list)
    for i, (_, norm, _) in enumerate(parsed):
        h = norm.get("hostname", "")
        if h:
            by_host[h].append(i)
    for idxs in by_host.values():
        if len(idxs) < 2:
            continue
        by_ip: dict[str, list[int]] = defaultdict(list)
        for i in idxs:
            by_ip[assets[i].primary_ip or ""].append(i)
        for ip, sub in by_ip.items():
            if ip and len(sub) > 1:
                for i in sub:
                    if assets[i].review_flag is None:
                        assets[i].review_flag = "DUP_HOSTNAME_SAMEIP"
                        review_breakdown["DUP_HOSTNAME_SAMEIP"] += 1


def parse_workbook(path: str) -> tuple[list[LedgerAsset], list[SheetStat]]:
    """전 보이는 시트 파싱 → (자산 목록, 시트별 통계). DB 미적재."""
    wb = load_workbook(path, read_only=True, data_only=True)
    by_norm = {ws.title.replace(" ", ""): ws for ws in wb.worksheets}

    all_assets: list[LedgerAsset] = []
    stats: list[SheetStat] = []
    for spec in SHEET_SPECS:
        ws = by_norm.get(spec.sheet_name.replace(" ", ""))
        if ws is None:
            logger.warning("시트 없음: %s", spec.sheet_name)
            continue
        assets, stat = parse_sheet(ws, spec)
        all_assets.extend(assets)
        stats.append(stat)
    return all_assets, stats
