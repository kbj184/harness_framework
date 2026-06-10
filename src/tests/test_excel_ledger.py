"""excel_ledger 파서·정규화·해싱 단위 테스트 (대외비 파일 비의존)."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from src.agents.excel_ledger import normalize as nz
from src.agents.excel_ledger.config import SHEET_SPECS
from src.agents.excel_ledger.hashing import asset_hash
from src.shared.cpe import cpe_for
from src.agents.excel_ledger.parser import HeaderMismatchError, parse_sheet


def _spec(cat):
    return next(s for s in SHEET_SPECS if s.category_cd == cat)


def _make_sheet(title, header_row, headers: dict[int, str], rows: list[dict[int, object]]):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for col, txt in headers.items():
        ws.cell(row=header_row, column=col, value=txt)
    for r, rowdata in enumerate(rows, start=header_row + 1):
        for col, val in rowdata.items():
            ws.cell(row=r, column=col, value=val)
    return ws


# ---------------- normalize ----------------

def test_account_key_aws_zero_pad():
    assert nz.account_key("aws_x@x.com(069682129442)") == "AWS:069682129442"
    # 보안시트 맨숫자(leading zero 제거)도 동일 12자리로 정규화
    assert nz.account_key("2715503011") == "AWS:002715503011"


def test_account_key_non_aws():
    assert nz.account_key("toast_dhs@gsretail.com") == "NHN_TOAST:toast_dhs@gsretail.com"
    assert nz.account_key("kakaocloud (gsretail-gsshop-lz)").startswith("KAKAO:")
    assert nz.account_key("NHN CLOUD").startswith("NHN_CLOUD:")


def test_ip_primary_parsing():
    assert nz.ip_primary("121.50.17.84 VIP : 121.50.17.83")[0] == "121.50.17.84"
    assert nz.ip_primary("N/A") == (None, [])
    assert nz.ip_primary("TBD '24.08.25") == (None, [])  # 4옥텟 IP 아님


def test_env_type():
    assert nz.env_type("하남센터 7F", "IDC") == "IDC"
    assert nz.env_type("ap-northeast-2", "CLOUD") == "CLOUD"
    assert nz.env_type("", "OA") == "OA"


# ---------------- hashing ----------------

def test_hash_deterministic_and_no_changes():
    h1 = asset_hash("HW_PC", ["pc1", "alice", "wg"])
    assert h1 == asset_hash("HW_PC", ["pc1", "alice", "wg"])
    assert asset_hash("HW_PC", ["pc1", "alice", "wg"], no="2") != h1  # NO 덧붙이면 달라짐


# ---------------- parser: PC 충돌 → NO surrogate + 플래그 ----------------

def test_pc_collision_no_surrogate():
    spec = _spec("HW_PC")
    headers = {5: "사용자 ID", 7: "컴퓨터 이름", 9: "MAC 주소"}
    rows = [
        {2: 1, 3: "O", 5: "alice", 7: "PC1", 8: "WG", 9: "AA:AA:AA:AA:AA:AA", 20: 1, 21: 1, 22: 1, 24: "하"},
        {2: 2, 3: "O", 5: "alice", 7: "PC1", 8: "WG", 9: "BB:BB:BB:BB:BB:BB", 20: 1, 21: 1, 22: 1, 24: "하"},
        {2: 3, 3: "O", 5: "bob", 7: "PC2", 8: "WG", 9: "CC:CC:CC:CC:CC:CC", 20: 1, 21: 1, 22: 1, 24: "하"},
    ]
    ws = _make_sheet("12. PC", 6, headers, rows)
    assets, stat = parse_sheet(ws, spec)

    assert stat.rows == 3
    assert stat.unique_hash == 3  # 충돌쌍도 NO로 유니크
    assert stat.review_queue == 2
    assert stat.review_breakdown["SUSPECTED_SAME_DEVICE"] == 2

    by_no = {a.source_id: a for a in assets}
    assert by_no["S12-1"].review_flag == "SUSPECTED_SAME_DEVICE"
    assert by_no["S12-1"].lifecycle_state == "CANDIDATE"
    assert by_no["S12-3"].review_flag is None
    assert by_no["S12-3"].lifecycle_state == "ACTIVE"
    assert len({a.asset_id_hash for a in assets}) == 3


# ---------------- parser: 클라우드 계정키 정규화 ----------------

def test_cloud_server_account_key_and_score():
    spec = _spec("CLD_SVR")
    headers = {8: "계정 이메일(ID)", 11: "instance-id", 13: "Private IP"}
    rows = [
        {2: 1, 3: "O", 8: "aws_x@x.com(069682129442)", 10: "FOO", 11: "i-0abc",
         13: "10.0.0.1", 26: 3, 27: 3, 28: 3, 30: "상"},
        {2: 2, 3: "O", 8: "toast_dhs@gsretail.com", 10: "BAR", 11: "i-0def",
         13: "10.0.0.2", 26: 3, 27: 3, 28: 3, 30: "상"},
    ]
    ws = _make_sheet("7. 클라우드 서버", 6, headers, rows)
    assets, stat = parse_sheet(ws, spec)

    assert stat.rows == 2 and stat.unique_hash == 2 and stat.review_queue == 0
    a0 = assets[0]
    assert a0.attributes["account_key"] == "AWS:069682129442"
    assert a0.attributes["account_provider"] == "AWS"
    assert a0.criticality_score == 9
    assert a0.env_type == "CLOUD"
    assert assets[1].attributes["account_provider"] == "NHN_TOAST"


# ---------------- parser: 서버 동명+동일IP 의미검토 ----------------

def test_server_dup_hostname_same_ip_flag():
    spec = _spec("HW_SVR")
    headers = {6: "자산명(hostname)", 8: "IP", 10: "모델명", 11: "OS"}
    rows = [
        # 동명+동일IP, 모델 상이 → 복합키 유니크지만 의미검토
        {2: 1, 3: "O", 6: "GSX01", 8: "10.0.0.5", 10: "ModelA", 11: "RHEL", 12: "7.9", 21: 1, 22: 1, 23: 1, 25: "하"},
        {2: 2, 3: "O", 6: "GSX01", 8: "10.0.0.5", 10: "ModelB", 11: "RHEL", 12: "7.9", 21: 1, 22: 1, 23: 1, 25: "하"},
        # 동명이지만 IP 상이 → 실물 2대, 플래그 없음
        {2: 3, 3: "O", 6: "GSY01", 8: "10.0.0.6", 10: "ModelC", 11: "AIX", 12: "7.2", 21: 1, 22: 1, 23: 1, 25: "하"},
        {2: 4, 3: "O", 6: "GSY01", 8: "10.0.0.7", 10: "ModelC", 11: "AIX", 12: "7.2", 21: 1, 22: 1, 23: 1, 25: "하"},
    ]
    ws = _make_sheet("2. 서버", 6, headers, rows)
    assets, stat = parse_sheet(ws, spec)

    by_no = {a.source_id: a for a in assets}
    assert by_no["S02-1"].review_flag == "DUP_HOSTNAME_SAMEIP"
    assert by_no["S02-2"].review_flag == "DUP_HOSTNAME_SAMEIP"
    assert by_no["S02-3"].review_flag is None  # IP 상이 = 실물 2대
    assert by_no["S02-4"].review_flag is None
    assert stat.unique_hash == 4  # 복합키(host+ip+model)로 전부 유니크


# ---------------- CPE 매핑 (D4) ----------------

def test_cpe_std_tomcat_prose_version():
    r = cpe_for("Tomcat", "Apache Tomcat Version 7.0.55")
    assert r.tier == "STD" and not r.unmapped
    assert r.vendor == "apache" and r.product == "tomcat" and r.version == "7.0.55"
    assert r.cpe_uri.startswith("cpe:2.3:a:apache:tomcat:7.0.55")


def test_cpe_os_part_o_and_aurora_upstream():
    assert cpe_for("RHEL", "7.9").cpe_uri.startswith("cpe:2.3:o:redhat:enterprise_linux:7.9")
    # aurora-postgresql → upstream postgresql
    r = cpe_for("aurora-postgresql", "16.6")
    assert (r.vendor, r.product, r.version) == ("postgresql", "postgresql", "16.6")


def test_cpe_manual_and_no_cpe_and_unknown():
    assert cpe_for("SmartZone", "6.1.1.0.959").tier == "MANUAL"
    lena = cpe_for("LENA", "lena 1.3")
    assert lena.tier == "NO_CPE" and lena.unmapped and lena.cpe_uri is None
    unk = cpe_for("듣보잡SW", "1.0")
    assert unk.tier == "UNKNOWN" and unk.unmapped


def test_cpe_windows_desktop_browsers():
    """Windows 설치 앱(msi) 중 CVE 빈발 브라우저 — CPE_SW 매칭 대상."""
    c = cpe_for("Chrome", "150.0.7850.0")
    assert c.tier == "STD" and (c.vendor, c.product) == ("google", "chrome")
    assert c.cpe_uri.startswith("cpe:2.3:a:google:chrome:150.0.7850.0")
    # Chromium 기반 Edge → edge_chromium (legacy edge 아님)
    e = cpe_for("Edge", "148.0.3967.96")
    assert (e.vendor, e.product) == ("microsoft", "edge_chromium")
    f = cpe_for("Firefox", "128.0")
    assert (f.vendor, f.product) == ("mozilla", "firefox")
    # OS 구성요소·AV 정의는 매핑 안 됨(UNKNOWN) → cpe_uri 안 채워져 매칭 제외
    assert cpe_for("Malware Protection", "1.451.193.0").tier == "UNKNOWN"
    assert cpe_for("Jscript", "5.812.10240.16384").tier == "UNKNOWN"


def test_cpe_appliance_os_mappings():
    """ISMS 시트5·6 실 OS값 → 정확한 벤더. 짧은 'ios' 오매칭(NIOS/FortiOS) 회귀 방지."""
    # 오매칭 버그 회귀 방지 — 'ios' 가 아닌 고유 벤더로
    assert cpe_for("NIOS").vendor == "infoblox"
    assert cpe_for("FortiOS").vendor == "fortinet"
    # 신규 추가 매핑
    asa = cpe_for("ASA OS")
    assert asa.vendor == "cisco" and asa.product == "adaptive_security_appliance_software"
    assert cpe_for("NOS").vendor == "brocade"            # Brocade VDX
    assert cpe_for("NS-OX").vendor == "citrix"           # Netscaler OS 변형
    # 국산·OEM → NO_CPE(unmapped, KISA行)
    assert cpe_for("Linux 4.14.128-2somansa2.el7").tier == "NO_CPE"
    assert cpe_for("OS-L3A-A").unmapped is True          # LG히타치 = ALAXALA OEM
    # 짧은 ios 는 여전히 동작
    assert cpe_for("IOS").product == "ios"


def test_parser_populates_cpe_and_unmapped_flag():
    spec = _spec("SW_WAS")
    headers = {5: "자산명(hostname)", 6: "APP 서비스", 9: "S/W"}
    rows = [
        {2: 1, 3: "o", 5: "WAS1", 6: "홈", 7: "운영", 8: "10.0.0.1", 9: "Tomcat", 10: "9.0.64",
         20: 3, 21: 3, 22: 3, 24: "상"},
        {2: 2, 3: "o", 5: "WAS2", 6: "GIS", 7: "운영", 8: "10.0.0.2", 9: "LENA", 10: "lena 1.3",
         20: 3, 21: 3, 22: 3, 24: "상"},
    ]
    ws = _make_sheet("3. WEBWAS", 6, headers, rows)
    assets, _ = parse_sheet(ws, spec)
    t = next(a for a in assets if a.source_id == "S03-1")
    assert t.cpe_product == "tomcat" and t.cpe_tier == "STD"
    lena = next(a for a in assets if a.source_id == "S03-2")
    assert lena.cpe_uri is None and lena.attributes.get("cpe_unmapped") is True


# ---------------- parser: 헤더 드리프트 감지 ----------------

def test_header_assert_catches_drift():
    spec = _spec("HW_PC")
    headers = {5: "엉뚱한헤더", 7: "컴퓨터 이름", 9: "MAC 주소"}  # col5 사용자 기대인데 틀림
    rows = [{2: 1, 5: "x", 7: "y", 8: "z", 9: "m"}]
    ws = _make_sheet("12. PC", 6, headers, rows)
    with pytest.raises(HeaderMismatchError):
        parse_sheet(ws, spec)
